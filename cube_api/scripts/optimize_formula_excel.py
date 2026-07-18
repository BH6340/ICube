import os
import openpyxl
from openpyxl import Workbook

EXCEL_PATH = r'E:\BH\PyStudy\web_projects\ICube\files\formula\CFOP_formula.xlsx'
OUTPUT_EXCEL_PATH = r'E:\BH\PyStudy\web_projects\ICube\files\formula\CFOP_formula_optimized.xlsx'
IMAGE_DIR = r'E:\BH\PyStudy\web_projects\ICube\cube_api\media\formulas'

NEW_HEADERS = [
    '序号', '公式名称', '公式记号', '分类-阶数', '分类-方法', 
    '分类-阶段', '难度等级', '公式描述', '缩略图文件名', '标签'
]

FORMULA_NAMES = {
    'F2L': {
        1: '基础槽位', 2: '角块在顶层', 3: '棱块在顶层', 4: '角块在底层',
        5: '棱块在底层', 6: '角块在中层', 7: '棱块在中层', 8: '角块在槽位',
        9: '棱块在槽位', 10: '角块被卡住', 11: '棱块被卡住', 12: '角块在侧面',
        13: '棱块在侧面', 14: '角块在背面', 15: '棱块在背面', 16: '角块在顶面',
        17: '棱块在顶面', 18: '角块在底面', 19: '棱块在底面', 20: '角块在左面',
        21: '棱块在左面', 22: '角块在右面', 23: '棱块在右面', 24: '角块在前面',
        25: '棱块在前面', 26: '角块在后面', 27: '棱块在后面', 28: '进阶槽位1',
        29: '进阶槽位2', 30: '进阶槽位3', 31: '进阶槽位4', 32: '进阶槽位5',
        33: '进阶槽位6', 34: '进阶槽位7', 35: '进阶槽位8', 36: '进阶槽位9',
        37: '进阶槽位10', 38: '进阶槽位11', 39: '进阶槽位12', 40: '进阶槽位13',
        41: '进阶槽位14'
    },
    'OLL': {
        1: '小鱼1', 2: '小鱼2', 3: '十字', 4: '一字', 5: '弯尺', 6: '点',
        7: '双凸', 8: '闪电', 9: '飞机', 10: '坦克', 11: '盾牌', 12: '双刺',
        13: '三角', 14: '双三角', 15: '风车', 16: '双风车', 17: '箭头',
        18: '双箭头', 19: '钩子', 20: '双钩', 21: '楼梯', 22: '双楼梯',
        23: 'U型', 24: '双U', 25: '蛇', 26: '双蛇', 27: '菱形', 28: '双菱形',
        29: '五角星', 30: '六角星', 31: '花', 32: '双花', 33: '月亮',
        34: '双月', 35: '太阳', 36: '双太阳', 37: '心形', 38: '双心',
        39: '蝴蝶', 40: '双蝴蝶', 41: '雪花', 42: '双雪花', 43: '皇冠',
        44: '双皇冠', 45: '钻石', 46: '双钻石', 47: '彩虹', 48: '双彩虹',
        49: '火焰', 50: '双火焰', 51: '水滴', 52: '双水滴', 53: '气泡',
        54: '双气泡', 55: '云朵', 56: '双云朵', 57: '流星'
    },
    'PLL': {
        1: '三棱换', 2: '三棱换2', 3: '四棱换', 4: '四棱换2', 5: '两角换',
        6: '两角换2', 7: '邻角对棱', 8: '邻角对棱2', 9: '对角对棱',
        10: '对角对棱2', 11: '邻角换', 12: '邻角换2', 13: '对角换',
        14: '对角换2', 15: '三角换', 16: '三角换2', 17: '四角换',
        18: '四角换2', 19: '三楞换', 20: '三楞换2', 21: '四楞换'
    }
}

DIFFICULTY_MAP = {
    'F2L': 3,
    'OLL': 3,
    'PLL': 4
}

DESCRIPTION_MAP = {
    'F2L': 'F2L基础公式，用于构建前两层槽位',
    'OLL': 'OLL公式，用于调整顶层朝向',
    'PLL': 'PLL公式，用于调整顶层位置'
}

TAG_MAP = {
    'F2L': 'F2L,基础',
    'OLL': 'OLL,常用',
    'PLL': 'PLL,进阶'
}


def rename_images():
    """重命名图片文件，统一命名格式"""
    for phase in ['F2L', 'OLL', 'PLL']:
        phase_dir = os.path.join(IMAGE_DIR, f'{phase}_Images')
        if not os.path.exists(phase_dir):
            continue
        
        for filename in sorted(os.listdir(phase_dir)):
            if filename.endswith('.png'):
                old_path = os.path.join(phase_dir, filename)
                
                if phase == 'F2L' and 'F2L_Images_' in filename:
                    new_name = filename.replace('F2L_Images_', 'F2L_')
                    new_path = os.path.join(phase_dir, new_name)
                    os.rename(old_path, new_path)
                    print(f'Renamed: {filename} -> {new_name}')
    
    print('图片重命名完成')


def optimize_excel():
    """优化Excel表格"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    new_wb = Workbook()
    
    for phase in ['F2L', 'OLL', 'PLL']:
        if phase not in wb.sheetnames:
            continue
        
        ws = wb[phase]
        new_ws = new_wb.create_sheet(title=phase)
        
        new_ws.append(NEW_HEADERS)
        
        for row_idx in range(2, ws.max_row + 1):
            old_row = list(ws[row_idx])
            seq = old_row[0].value
            notation = old_row[1].value
            old_image_name = old_row[2].value if len(old_row) > 2 else None
            
            formula_name = FORMULA_NAMES.get(phase, {}).get(seq, f'{phase}{seq}')
            difficulty = DIFFICULTY_MAP.get(phase, 3)
            description = DESCRIPTION_MAP.get(phase, '')
            tags = TAG_MAP.get(phase, '')
            
            if old_image_name:
                new_image_name = old_image_name
                if phase == 'F2L' and 'F2L_Images_' in old_image_name:
                    new_image_name = old_image_name.replace('F2L_Images_', 'F2L_')
            else:
                new_image_name = f'{phase}_{str(seq).zfill(3)}.png'
            
            new_row = [
                seq, formula_name, notation, 3, 'CFOP', phase,
                difficulty, description, new_image_name, tags
            ]
            new_ws.append(new_row)
    
    if 'Sheet' in new_wb.sheetnames:
        del new_wb['Sheet']
    
    new_wb.save(OUTPUT_EXCEL_PATH)
    print(f'优化后的Excel已保存到: {OUTPUT_EXCEL_PATH}')


if __name__ == '__main__':
    rename_images()
    optimize_excel()
    print('所有操作完成！')