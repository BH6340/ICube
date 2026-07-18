import os
import openpyxl
from django.core.management.base import BaseCommand
from django.conf import settings

from apps.formula.models import CubeCategory, Formula, FormulaTag, FormulaTagRelation
from apps.formula.services import FormulaService


EXCEL_PATH = r'E:\BH\PyStudy\web_projects\ICube\files\formula\CFOP_formula_optimized.xlsx'
IMAGE_BASE_DIR = r'E:\BH\PyStudy\web_projects\ICube\cube_api\media\formulas'


class Command(BaseCommand):
    help = '导入 CFOP 公式数据'

    def handle(self, *args, **options):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        
        total_imported = 0
        total_categories = 0
        total_tags = 0

        for phase in ['F2L', 'OLL', 'PLL']:
            if phase not in wb.sheetnames:
                self.stdout.write(f'Sheet {phase} 不存在，跳过')
                continue

            ws = wb[phase]
            self.stdout.write(f'正在导入 {phase} 公式...')

            category, created = CubeCategory.objects.get_or_create(
                order=3,
                method='CFOP',
                phase=phase,
                defaults={
                    'name': f'三阶魔方 CFOP 法 {phase}',
                    'description': f'三阶魔方 CFOP 方法的 {phase} 阶段',
                    'sort_order': {'F2L': 1, 'OLL': 2, 'PLL': 3}[phase]
                }
            )
            if created:
                total_categories += 1
                self.stdout.write(f'创建分类: {category.name}')

            for row_idx in range(2, ws.max_row + 1):
                row = list(ws[row_idx])
                
                seq = row[0].value
                name = row[1].value
                notation = row[2].value
                order = row[3].value
                method = row[4].value
                phase_name = row[5].value
                difficulty = row[6].value
                description = row[7].value
                image_filename = row[8].value
                tags_str = row[9].value if len(row) > 9 else None

                if not notation:
                    self.stdout.write(f'第 {row_idx} 行公式记号为空，跳过')
                    continue

                image_relative_path = None
                if image_filename:
                    image_subdir = f'{phase}_Images'
                    image_path = os.path.join(IMAGE_BASE_DIR, image_subdir, image_filename)
                    if os.path.exists(image_path):
                        image_relative_path = f'formulas/{image_subdir}/{image_filename}'
                    else:
                        self.stdout.write(f'图片不存在: {image_path}')

                inverse_notation = FormulaService.generate_inverse_notation(notation)

                formula, created = Formula.objects.get_or_create(
                    category=category,
                    name=name,
                    defaults={
                        'notation': notation,
                        'inverse_notation': inverse_notation,
                        'difficulty': difficulty,
                        'description': description,
                        'thumbnail': image_relative_path,
                        'is_custom': False
                    }
                )

                if created:
                    total_imported += 1
                    self.stdout.write(f'创建公式: {name}')
                else:
                    formula.notation = notation
                    formula.inverse_notation = inverse_notation
                    formula.difficulty = difficulty
                    formula.description = description
                    if image_relative_path:
                        formula.thumbnail = image_relative_path
                    formula.save()
                    self.stdout.write(f'更新公式: {name}')

                if tags_str:
                    tags = [t.strip() for t in tags_str.split(',') if t.strip()]
                    for tag_name in tags:
                        tag, tag_created = FormulaTag.objects.get_or_create(name=tag_name)
                        if tag_created:
                            total_tags += 1

                        FormulaTagRelation.objects.get_or_create(
                            formula=formula,
                            tag=tag
                        )

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(f'导入完成！'))
        self.stdout.write(f'分类: 创建 {total_categories} 个')
        self.stdout.write(f'公式: 导入 {total_imported} 个')
        self.stdout.write(f'标签: 创建 {total_tags} 个')